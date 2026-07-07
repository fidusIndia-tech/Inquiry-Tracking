"""
attachment_handler.py
---------------------
Downloads attachments from Gmail and extracts text from them.

Handles:
  - PDF  → text via pdfplumber
  - XLSX / XLS → text via openpyxl / xlrd
  - DOCX → text via python-docx
  - Plain .txt / .csv → direct decode

Workers call extract_attachment_text(service, message_id) to get
all attachment text as a single concatenated string.
"""

import base64
import io
from logging_setup import get_logger

logger = get_logger(__name__)


def _download_attachment(service, message_id: str, attachment_id: str) -> bytes:
    """Download raw attachment bytes from Gmail API."""
    result = (
        service.users()
        .messages()
        .attachments()
        .get(userId="me", messageId=message_id, id=attachment_id)
        .execute()
    )
    data = result.get("data", "")
    return base64.urlsafe_b64decode(data + "==")


def _ocr_pdf(data: bytes) -> str:
    """OCR fallback for scanned/image-only PDFs. Requires: pip install pdf2image pytesseract
    and system packages: poppler-utils + tesseract-ocr."""
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
        images = convert_from_bytes(data, dpi=200)
        pages = [pytesseract.image_to_string(img) for img in images]
        return "\n".join(pages)
    except ImportError:
        logger.debug("pdf2image/pytesseract not installed — OCR unavailable. pip install pdf2image pytesseract (+ system: poppler-utils tesseract-ocr)")
        return ""
    except Exception as exc:
        logger.warning("PDF OCR failed: %s", exc)
        return ""


def _extract_from_pdf(data: bytes) -> str:
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            pages = [p.extract_text() or "" for p in pdf.pages]
        text = "\n".join(pages)
    except ImportError:
        logger.warning("pdfplumber not installed — skipping PDF. Run: pip install pdfplumber")
        return ""
    except Exception as exc:
        logger.warning("PDF extraction failed: %s", exc)
        return ""

    # Scanned/image-based PDF: pdfplumber returns near-empty text. Try OCR.
    if len(text.strip()) < 50:
        logger.info("PDF text minimal (%d chars) — attempting OCR", len(text.strip()))
        ocr_text = _ocr_pdf(data)
        if ocr_text.strip():
            return ocr_text
    return text


def _extract_from_xlsx(data: bytes) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        rows = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    rows.append("\t".join(cells))
        return "\n".join(rows)
    except ImportError:
        logger.warning("openpyxl not installed — skipping XLSX. Run: pip install openpyxl")
        return ""
    except Exception as exc:
        logger.warning("XLSX extraction failed: %s", exc)
        return ""


def _extract_from_docx(data: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        logger.warning("python-docx not installed — skipping DOCX. Run: pip install python-docx")
        return ""
    except Exception as exc:
        logger.warning("DOCX extraction failed: %s", exc)
        return ""


def _extract_from_ods(data: bytes) -> str:
    """Extract text from OpenDocument Spreadsheet (.ods). Requires: pip install odfpy"""
    try:
        from odf.opendocument import load as odf_load
        from odf.table import Table, TableRow, TableCell
        from odf import teletype
        doc = odf_load(io.BytesIO(data))
        rows = []
        for sheet in doc.spreadsheet.getElementsByType(Table):
            for row in sheet.getElementsByType(TableRow):
                cells = [
                    teletype.extractText(cell).strip()
                    for cell in row.getElementsByType(TableCell)
                ]
                if any(cells):
                    rows.append("\t".join(cells))
        return "\n".join(rows)
    except ImportError:
        logger.warning("odfpy not installed — skipping ODS. Run: pip install odfpy")
        return ""
    except Exception as exc:
        logger.warning("ODS extraction failed: %s", exc)
        return ""


def _extract_from_zip(data: bytes) -> str:
    """Unzip and extract text from all recognised file types inside the archive."""
    import zipfile
    try:
        texts = []
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for name in zf.namelist():
                if name.endswith("/"):
                    continue
                ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
                try:
                    raw = zf.read(name)
                except Exception as exc:
                    logger.warning("ZIP: failed to read %s: %s", name, exc)
                    continue
                if ext == "pdf":
                    text = _extract_from_pdf(raw)
                elif ext in ("xlsx", "xls"):
                    text = _extract_from_xlsx(raw)
                elif ext in ("docx", "doc"):
                    text = _extract_from_docx(raw)
                elif ext == "ods":
                    text = _extract_from_ods(raw)
                elif ext in ("txt", "csv"):
                    text = raw.decode("utf-8", errors="replace")
                else:
                    text = ""
                if text.strip():
                    texts.append(f"[ZIP/{name}]\n{text.strip()}")
        return "\n\n".join(texts)
    except Exception as exc:
        logger.warning("ZIP extraction failed: %s", exc)
        return ""


_SPREADSHEET_EXTS = {"xlsx", "xls", "csv"}


def _normalize_hdr(h) -> str:
    """Lowercase + collapse whitespace for flexible header matching."""
    return " ".join(str(h or "").lower().strip().split())


def _map_header(raw: str) -> str | None:
    """Return canonical field name for a column header, or None to ignore."""
    h = _normalize_hdr(raw)
    if h in ("part no", "part no.", "part number", "item code", "material code",
             "item no", "item no.", "part nos", "part nos.", "material no",
             "material no."):
        return "part_number"
    if "qty" in h or "quantity" in h:
        return "quantity"
    if "short description" in h or h == "short desc":
        return "short_desc"
    if h in ("part description", "description", "product description",
             "item description", "material description", "long description"):
        return "description"
    if h in ("purch u/m", "uom", "unit", "unit of measure", "u/m", "u.o.m",
             "purch uom", "purchase uom", "unit of msmt"):
        return "uom"
    if h in ("brand", "make", "manufacturer", "oem", "make / brand", "brand/make",
             "make/brand"):
        return "brand"
    if "lead time" in h or "lead_time" in h or "delivery time" in h:
        return "lead_time"
    if h in ("remarks", "remark", "notes", "note", "comments", "comment"):
        return "remarks"
    # Rate/HSN/GST — capture for notes but not core fields
    if "rate" in h or "price" in h:
        return "rate"
    if "hsn" in h or "sac" in h:
        return "hsn"
    if "gst" in h or ("tax" in h and "%" in h):
        return "gst"
    return None


def find_spreadsheet_attachment(payload: dict) -> tuple[str, str] | None:
    """
    Peek at the MIME payload tree (no download) and return (filename,
    attachment_id) for the first xlsx/xls/csv attachment found, or None.
    """
    result: list[tuple[str, str]] = []

    def walk(part: dict) -> None:
        filename = part.get("filename", "")
        body = part.get("body", {})
        attachment_id = body.get("attachmentId")
        if filename and attachment_id:
            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            if ext in _SPREADSHEET_EXTS:
                result.append((filename, attachment_id))
        for sub in part.get("parts", []):
            walk(sub)

    walk(payload)
    return result[0] if result else None


def parse_excel_rfq(data: bytes, filename: str = "") -> list[dict]:
    """
    Deterministically parse an xlsx/xls/csv attachment into structured RFQ
    item dicts.  No GPT involved — no token limits, no truncation.

    Returns [] if the sheet has no recognisable RFQ headers (so the caller
    falls through to the normal email body / GPT path).
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "xlsx"
    rows: list[tuple] = []

    try:
        if ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.worksheets[0]
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
            wb.close()
        elif ext == "csv":
            import csv as _csv
            text = data.decode("utf-8", errors="replace")
            rows = [tuple(r) for r in _csv.reader(text.splitlines())]
    except ImportError as exc:
        logger.warning("Excel RFQ parse: missing library — %s", exc)
        return []
    except Exception as exc:
        logger.warning("Excel RFQ parse failed for %s: %s", filename, exc)
        return []

    if not rows:
        return []

    # Find the header row (first row that contains at least 2 mapped columns).
    header_idx = None
    col_map: dict[int, str] = {}
    for ridx, row in enumerate(rows[:10]):  # headers are always in first 10 rows
        mapped = {}
        for cidx, cell in enumerate(row):
            field = _map_header(str(cell or ""))
            if field:
                mapped[cidx] = field
        if len(mapped) >= 2:
            header_idx = ridx
            col_map = mapped
            break

    if header_idx is None:
        logger.info("Excel RFQ: no recognisable RFQ headers in %s — skipping", filename)
        return []

    # Must have at least a part_number or description column to be useful.
    if not any(f in col_map.values() for f in ("part_number", "description", "short_desc")):
        logger.info("Excel RFQ: no part/description column in %s — skipping", filename)
        return []

    logger.info(
        "Excel RFQ: detected sheet in %s | columns=%s",
        filename,
        {v: k for k, v in col_map.items()},
    )

    items: list[dict] = []
    for row in rows[header_idx + 1:]:
        def cell(field: str) -> str:
            for cidx, f in col_map.items():
                if f == field and cidx < len(row):
                    v = row[cidx]
                    return str(v).strip() if v is not None else ""
            return ""

        part_number = cell("part_number")
        short_desc  = cell("short_desc")
        description = cell("description") or short_desc
        brand       = cell("brand")
        uom         = cell("uom")
        lead_time   = cell("lead_time")
        remarks     = cell("remarks")
        rate        = cell("rate")
        hsn         = cell("hsn")
        gst         = cell("gst")

        # Parse quantity — keep as string if non-numeric
        qty_raw = cell("quantity")
        try:
            quantity: str | int | float = int(float(qty_raw)) if qty_raw else ""
        except (ValueError, TypeError):
            quantity = qty_raw

        # Skip completely empty rows
        if not any([part_number, description, short_desc]):
            continue

        # Build notes from all descriptive fields so nothing is lost
        note_parts = []
        if description and description != short_desc:
            note_parts.append(description)
        elif short_desc:
            note_parts.append(short_desc)
        if lead_time:
            note_parts.append(f"Lead time: {lead_time}")
        if remarks:
            note_parts.append(f"Remarks: {remarks}")
        if rate:
            note_parts.append(f"Rate: {rate}")
        if hsn:
            note_parts.append(f"HSN/SAC: {hsn}")
        if gst:
            note_parts.append(f"GST: {gst}%")

        items.append({
            "part_number": part_number or None,
            "brand":       brand or "",
            "quantity":    quantity,
            "uom":         uom or "",
            "notes":       " | ".join(note_parts) if note_parts else (short_desc or description),
        })

    logger.info("Excel RFQ: extracted %d items from %s", len(items), filename)
    return items


def download_and_parse_excel_rfq(
    service, message_id: str, payload: dict
) -> list[dict]:
    """
    One-call helper used by the email pipeline: find the first spreadsheet
    attachment, download it, and parse it as a structured RFQ.
    Returns [] if no spreadsheet found or no RFQ headers detected.
    """
    info = find_spreadsheet_attachment(payload)
    if not info:
        return []
    filename, attachment_id = info
    logger.info("Excel RFQ: found attachment %s — downloading", filename)
    try:
        raw = _download_attachment(service, message_id, attachment_id)
    except Exception as exc:
        logger.warning("Excel RFQ: download failed for %s: %s", filename, exc)
        return []
    return parse_excel_rfq(raw, filename)


def extract_attachment_text(service, message_id: str, payload: dict) -> str:
    """
    Walk the message payload tree, find all attachments,
    download and extract text from each.
    Returns concatenated text from all attachments.
    """
    texts = []

    def walk(part: dict):
        filename = part.get("filename", "")
        body     = part.get("body", {})
        attachment_id = body.get("attachmentId")
        inline_data   = body.get("data")  # small files Gmail inlines in body.data

        if filename and (attachment_id or inline_data):
            mime = part.get("mimeType", "")
            # Skip text/multipart inline parts — those are the email body, not files
            if not attachment_id and mime.startswith(("text/", "multipart/")):
                for subpart in part.get("parts", []):
                    walk(subpart)
                return

            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            logger.info("  → extracting attachment: %s (%s)", filename, ext)

            try:
                if attachment_id:
                    raw = _download_attachment(service, message_id, attachment_id)
                else:
                    import base64 as _b64
                    raw = _b64.urlsafe_b64decode(inline_data + "==")
            except Exception as exc:
                logger.warning("  → download failed for %s: %s", filename, exc)
                return

            if ext == "pdf":
                text = _extract_from_pdf(raw)
            elif ext in ("xlsx", "xls"):
                text = _extract_from_xlsx(raw)
            elif ext in ("docx", "doc"):
                text = _extract_from_docx(raw)
            elif ext == "ods":
                text = _extract_from_ods(raw)
            elif ext == "zip":
                text = _extract_from_zip(raw)
            elif ext in ("txt", "csv"):
                text = raw.decode("utf-8", errors="replace")
            else:
                logger.debug("  → unsupported attachment type: %s", ext)
                text = ""

            if text.strip():
                texts.append(f"[Attachment: {filename}]\n{text.strip()}")

        for subpart in part.get("parts", []):
            walk(subpart)

    walk(payload)
    return "\n\n".join(texts)
